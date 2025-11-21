# scalanie.py
# Łączy wszystkie wojewódzkie CSV z jednego folderu w JEDEN arkusz Excela
# o nazwie: "Polska (HH.MM dd.mm.RRRR)" (limit 31 znaków – Excel).
#
# Użycie:
#   python scalanie.py --input <folder_z_csv> --output <plik_wyjściowy.xlsx>
# Opcjonalnie:
#   --pattern "*.csv"  (domyślnie)
#   --encoding "utf-8-sig" (domyślnie)
#   --sort             (posortuje po wojewodztwo, miejscowosc, dzielnica)

from __future__ import annotations
import argparse
from pathlib import Path
from datetime import datetime
from zoneinfo import ZoneInfo
import re
import sys
import csv

import pandas as pd

# Główne kolumny z ogłoszeń
HEADERS = [
    "cena","cena_za_metr","metry","liczba_pokoi","pietro","rynek","rok_budowy","material",
    "wojewodztwo","powiat","gmina","miejscowosc","dzielnica","ulica","link"
]

# Dodatkowa kolumna na błędy parsowania pojedynczych linii
ERROR_COL = "blad_parsowania"

INVALID_SHEET_CHARS = r'[\[\]\*\?\/\\:]'   # Excel sheet name invalid chars


def log(msg: str) -> None:
    print(msg, flush=True)


def safe_sheet_name(name: str) -> str:
    name = re.sub(INVALID_SHEET_CHARS, "_", name)
    # Excel limit 31
    return name[:31] if len(name) > 31 else name


def _merge_decimal_fields(fields: list[str]) -> list[str]:
    """
    Jeśli wiersz ma więcej pól niż HEADERS (15), próbujemy naprawić:
    1) Założyć, że 'cena' została rozbita na 2 pola (indeks 0 i 1)
    2) Jeśli nadal za dużo pól, próbujemy skleić 'metry' (indeks 2 i 3)
    Jeśli po tych krokach wciąż liczba pól != 15, zwracamy listę tak jak jest
    (wyżej oznaczymy to jako błąd).
    """
    # 1) sklej potencjalnie rozbitą cenę
    if len(fields) > len(HEADERS):
        # logika: pierwsze pole wygląda jak część liczby, drugie jak 'grosze'
        f0 = fields[0].strip().replace('"', '').replace(" ", "")
        f1 = fields[1].strip().replace('"', '').replace(" ", "") if len(fields) > 1 else ""
        if f0.replace("0", "1").isdigit() and f1.isdigit() and len(f1) <= 2:
            fields[0] = fields[0] + "," + fields[1]
            del fields[1]

    # 2) sklej potencjalnie rozbite metry
    if len(fields) > len(HEADERS) and len(fields) >= 4:
        f2 = fields[2].strip().replace('"', '').replace(" ", "")
        f3 = fields[3].strip().replace('"', '').replace(" ", "") if len(fields) > 3 else ""
        if f2.replace("0", "1").isdigit() and f3.isdigit() and len(f3) <= 2:
            fields[2] = fields[2] + "," + fields[3]
            del fields[3]

    # jeśli dalej jest za dużo pól – zostaw, wyżej oznaczymy jako błąd
    return fields


def read_csvs(in_dir: Path, pattern: str, encoding: str) -> list[pd.DataFrame]:
    files = sorted(in_dir.glob(pattern))
    if not files:
        log(f"[WARN] Brak plików pasujących do wzorca: {pattern} w {in_dir}")
        return []

    dfs: list[pd.DataFrame] = []

    for f in files:
        log(f"[READ] {f.name}")
        rows: list[dict] = []
        try:
            with f.open("r", encoding=encoding, newline="") as fh:
                reader = csv.reader(fh, delimiter=";", quoting=csv.QUOTE_MINIMAL)

                for line_no, fields in enumerate(reader, start=1):
                    if not fields:
                        continue

                    # Pomijamy wiersz nagłówka (cena; cena_za_metr; ...)
                    if line_no == 1 and fields[0].strip().lower() == "cena":
                        continue

                    # Wiersz typu "ERROR: ..." – zapisz w kolumnie błędu
                    if len(fields) == 1 and fields[0].startswith("ERROR"):
                        rows.append({
                            **{h: "" for h in HEADERS},
                            ERROR_COL: fields[0]
                        })
                        continue

                    # Dopasuj liczbę pól
                    fields = _merge_decimal_fields(fields)

                    if len(fields) != len(HEADERS):
                        # Nie udało się dopasować – wszystko wrzucamy w blad_parsowania
                        raw = ";".join(fields)
                        rows.append({
                            **{h: "" for h in HEADERS},
                            ERROR_COL: f"LINIA {line_no}: {raw}"
                        })
                        continue

                    # OK – zwykły wiersz ogłoszenia
                    row = dict(zip(HEADERS, fields))
                    row[ERROR_COL] = ""
                    rows.append(row)

            if not rows:
                log(f"[WARN] Plik {f.name} nie zawiera poprawnych wierszy.")
                continue

            df = pd.DataFrame(rows, columns=HEADERS + [ERROR_COL])

            # Uzupełnij puste województwo nazwą pliku (tak jak wcześniej)
            woj = f.stem.lower().replace(".__tmp__", "")
            mask = df["wojewodztwo"].astype(str).str.strip().eq("")
            if mask.any():
                df.loc[mask, "wojewodztwo"] = woj

            dfs.append(df)

        except Exception as e:
            log(f"[ERR] Nie udało się wczytać {f}: {e}")

    return dfs


def autosize_columns(ws) -> None:
    """Proste auto-dopasowanie szerokości kolumn (openpyxl Worksheet)."""
    from openpyxl.utils import get_column_letter

    # Zbierz maksymalną długość tekstu dla każdej kolumny
    for i, col in enumerate(ws.iter_cols(1, ws.max_column), start=1):
        max_len = 0
        for cell in col:
            try:
                v = cell.value
                if v is None:
                    l = 0
                else:
                    l = len(str(v))
                if l > max_len:
                    max_len = l
            except Exception:
                pass
        # Minimalnie 8, maks 60, dodaj mały margines
        width = max(8, min(60, max_len + 2))
        ws.column_dimensions[get_column_letter(i)].width = width


def write_excel(df: pd.DataFrame, out_xlsx: Path) -> None:
    # Nazwa arkusza z godziną i datą w strefie Europe/Warsaw
    now = datetime.now(ZoneInfo("Europe/Warsaw"))
    stamp = now.strftime("%H.%M %d.%m.%Y")    # UWAGA: kropka zamiast ":" (Excel nie lubi ":")
    sheet_name = safe_sheet_name(f"Polska ({stamp})")

    log(f"[WRITE] {out_xlsx.name}  arkusz='{sheet_name}'  wierszy={len(df)}")
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)

    with pd.ExcelWriter(out_xlsx, engine="openpyxl", mode="w") as writer:
        df.to_excel(writer, index=False, sheet_name=sheet_name)

        # Formatowanie arkusza (autofilter, freeze, autosize)
        ws = writer.book[sheet_name]
        # Freeze top row
        ws.freeze_panes = "A2"
        # Autofilter na cały zakres danych
        ws.auto_filter.ref = ws.dimensions
        # Szerokości kolumn
        autosize_columns(ws)


def main():
    ap = argparse.ArgumentParser(
        description="Scalanie wojewódzkich CSV do Excela (1 arkusz: Polska (HH.MM dd.mm.RRRR))"
    )
    ap.add_argument("--input", required=True, help="Folder z plikami CSV (województwa).")
    ap.add_argument("--output", required=True, help="Ścieżka wyjściowa do pliku .xlsx.")
    ap.add_argument("--pattern", default="*.csv", help="Wzorzec plików (domyślnie: *.csv).")
    ap.add_argument("--encoding", default="utf-8-sig", help="Kodowanie CSV (domyślnie: utf-8-sig).")
    ap.add_argument("--sort", action="store_true",
                    help="Sortuj po (wojewodztwo, miejscowosc, dzielnica).")
    args = ap.parse_args()

    in_dir = Path(args.input)
    out_xlsx = Path(args.output)

    if not in_dir.exists() or not in_dir.is_dir():
        log(f"[ERR] Katalog wejściowy nie istnieje lub nie jest katalogiem: {in_dir}")
        sys.exit(2)

    log(f"[START] scalanie z: {in_dir}  ->  {out_xlsx}")
    dfs = read_csvs(in_dir, args.pattern, args.encoding)
    if not dfs:
        log("[ERR] Nie znaleziono żadnych danych do scalenia.")
        sys.exit(1)

    # Konkatenacja
    df = pd.concat(dfs, ignore_index=True)

    # Dedup po linku (zachowaj pierwsze wystąpienie)
    if "link" in df.columns:
        before = len(df)
        df = df.drop_duplicates(subset=["link"], keep="first")
        log(f"[DEDUP] link: {before} -> {len(df)}")

    # Opcjonalne sortowanie
    if args.sort:
        for col in ("wojewodztwo", "miejscowosc", "dzielnica"):
            if col not in df.columns:
                df[col] = ""
        df = df.sort_values(
            ["wojewodztwo", "miejscowosc", "dzielnica"],
            kind="stable",
            ignore_index=True
        )

    # Upewnij się, że mamy komplet kolumn: HEADERS + ERROR_COL
    for col in HEADERS:
        if col not in df.columns:
            df[col] = ""
    if ERROR_COL not in df.columns:
        df[ERROR_COL] = ""

    df = df[HEADERS + [ERROR_COL]]

    write_excel(df, out_xlsx)
    log("[DONE] Zapisano plik.")


if __name__ == "__main__":
    main()
