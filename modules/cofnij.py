#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from pathlib import Path
import argparse
from typing import List, Dict
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_SOURCE_SHEET = "raport_odfiltrowane"


def _norm_header(s: str | None) -> str:
    return (str(s or "")
            .replace("\u00a0", " ")
            .strip()
            .casefold())


def _read_header(ws: Worksheet) -> List[str]:
    max_c = ws.max_column
    header = [ws.cell(row=1, column=c).value for c in range(1, max_c + 1)]
    # obetnij puste ogony
    while header and (header[-1] is None or str(header[-1]).strip() == ""):
        header.pop()
    return [str(h) if h is not None else "" for h in header]


def _row_has_data(ws: Worksheet, row: int, max_col: int) -> bool:
    for c in range(1, max_col + 1):
        v = ws.cell(row=row, column=c).value
        if v not in (None, ""):
            return True
    return False


def _last_filled_row(ws: Worksheet, max_col: int) -> int:
    r = ws.max_row
    while r > 1:
        if _row_has_data(ws, r, max_col):
            return r
        r -= 1
    # jeśli tylko nagłówek jest wypełniony, to "ostatni wypełniony" = 1
    return 1 if _row_has_data(ws, 1, max_col) else 0


def _pick_target_sheet_name(wb, preferred: str | None, source_name: str) -> str:
    if preferred and preferred in wb.sheetnames:
        return preferred
    # domyślnie: pierwszy arkusz różny od źródłowego
    for name in wb.sheetnames:
        if name != source_name:
            return name
    # fallback: jeśli został tylko arkusz źródłowy — użyj go (dokleimy tam)
    return wb.sheetnames[0]


def main():
    ap = argparse.ArgumentParser(description="Cofnij wiersze z 'raport_odfiltrowane' do arkusza raportu.")
    ap.add_argument("--in", dest="input_path", required=True, help="Ścieżka do pliku Excel")
    ap.add_argument("--sheet", dest="target_sheet", default=None, help="Nazwa arkusza docelowego (raport)")
    ap.add_argument("--source", dest="source_sheet", default=DEFAULT_SOURCE_SHEET, help="Nazwa arkusza źródłowego (domyślnie 'raport_odfiltrowane')")
    ap.add_argument("--no-clear", dest="no_clear", action="store_true", help="Nie czyść danych w arkuszu źródłowym po przeniesieniu")
    args = ap.parse_args()

    xlsx = Path(args.input_path).expanduser()
    if not xlsx.exists():
        print(f"[ERR] Plik nie istnieje: {xlsx}")
        raise SystemExit(1)

    try:
        wb = load_workbook(xlsx)
    except Exception as e:
        print(f"[ERR] Nie mogę wczytać skoroszytu: {e}")
        raise SystemExit(2)

    if args.source_sheet not in wb.sheetnames:
        print(f"[ERR] Brak arkusza źródłowego: {args.source_sheet}")
        raise SystemExit(3)

    ws_src: Worksheet = wb[args.source_sheet]
    # wybierz arkusz docelowy
    tgt_name = _pick_target_sheet_name(wb, args.target_sheet, args.source_sheet)
    ws_tgt: Worksheet = wb[tgt_name]

    # nagłówki
    src_header = _read_header(ws_src)
    tgt_header = _read_header(ws_tgt)

    if not src_header:
        print("[INFO] Arkusz źródłowy nie ma nagłówka lub jest pusty – nic do zrobienia.")
        raise SystemExit(0)
    if not tgt_header:
        # jeśli w docelowym nie ma nagłówka – skopiuj nagłówek źródła
        for ci, name in enumerate(src_header, start=1):
            ws_tgt.cell(row=1, column=ci, value=name)
        tgt_header = src_header[:]

    # mapowanie nagłówków po znormalizowanej nazwie
    src_map: Dict[str, int] = {_norm_header(h): i for i, h in enumerate(src_header)}  # 0-based index
    tgt_norm = [_norm_header(h) for h in tgt_header]

    # policz niepuste wiersze źródła
    max_col_src = max(len(src_header), ws_src.max_column)
    # start danych źródłowych od 2. wiersza
    rows_to_copy = []
    for r in range(2, ws_src.max_row + 1):
        if _row_has_data(ws_src, r, max_col_src):
            # zbierz wartości w kolejności nagłówka źródła (do listy 0..len-1)
            row_vals = [ws_src.cell(row=r, column=c).value for c in range(1, len(src_header) + 1)]
            rows_to_copy.append(row_vals)

    if not rows_to_copy:
        print("[INFO] Brak danych do przeniesienia w arkuszu źródłowym.")
        raise SystemExit(0)

    # wyznacz miejsce doklejenia w docelowym
    max_col_tgt = max(len(tgt_header), ws_tgt.max_column)
    last_filled = _last_filled_row(ws_tgt, max_col_tgt)
    start_row = max(last_filled + 1, 2)  # zawsze poniżej nagłówka

    # wstawianie z dopasowaniem kolumn po nazwach
    written = 0
    for rv in rows_to_copy:
        # zmapuj: wartości w kolejności nagłówków docelowych
        out_row = []
        for norm_name in tgt_norm:
            idx = src_map.get(norm_name, None)
            out_row.append(rv[idx] if idx is not None and idx < len(rv) else "")
        # zapisz
        r = start_row + written
        for c_idx, val in enumerate(out_row, start=1):
            ws_tgt.cell(row=r, column=c_idx, value=val)
        written += 1

    # po przeniesieniu – czyść dane w źródle (od 2. wiersza w dół), o ile nie wyłączono
    if not args.no_clear:
        if ws_src.max_row > 1:
            ws_src.delete_rows(2, ws_src.max_row - 1)

    try:
        wb.save(xlsx)
    except PermissionError:
        print("[ERR] Nie mogę zapisać – plik jest otwarty w Excelu. Zamknij go i spróbuj ponownie.")
        raise SystemExit(4)
    except Exception as e:
        print(f"[ERR] Błąd zapisu: {e}")
        raise SystemExit(5)

    print(f"[OK] Przeniesiono {written} wierszy z „{args.source_sheet}” do „{tgt_name}” pod wiersz {last_filled}.")
    if not args.no_clear:
        print(f"[OK] Wyczyściłem dane w „{args.source_sheet}” (pozostawiono nagłówek).")
    else:
        print("[INFO] Dane w arkuszu źródłowym pozostawiono bez zmian (--no-clear).")


if __name__ == "__main__":
    main()
