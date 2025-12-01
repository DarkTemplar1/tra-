#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import argparse
import re
import unicodedata
from pathlib import Path
from typing import Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ----------------------------- utils -----------------------------

def desktop_dir() -> Path:
    return Path.home() / "Desktop"

def baza_dir() -> Path:
    return desktop_dir() / "baza danych"

def resolve_default(path_like: Optional[str], fallback_name: str) -> Path:
    if path_like:
        p = Path(path_like)
        if p.exists():
            return p
    return baza_dir() / fallback_name

def norm_colnames(columns: List[str]) -> Dict[str, str]:
    return {str(c).strip().lower(): c for c in columns}

def ensure_columns(df: pd.DataFrame, required: List[str]) -> Dict[str, str]:
    cmap = norm_colnames(df.columns)
    for c in required:
        if c not in cmap:
            df[c] = pd.Series([pd.NA] * len(df), dtype="object")
            cmap[c] = c
    return cmap

def norm_text(s: object) -> Optional[str]:
    if pd.isna(s):
        return None
    t = str(s).strip().lower()
    t = "".join(ch for ch in unicodedata.normalize("NFKD", t) if not unicodedata.combining(ch))
    t = re.sub(r"\s+", " ", t)
    t = t.replace(" - ", "-").replace(".", "")
    return t or None

def extract_street_from_link(link: object) -> Optional[str]:
    if pd.isna(link):
        return None
    s = str(link)
    m = re.search(r"(?:^|[^\w])ul\.\s*([A-ZĄĆĘŁŃÓŚŹŻa-ząćęłńóśźż0-9 .\-']+)", s)
    if not m:
        return None
    street = m.group(1)
    street = re.split(r"[/?#]", street)[0].strip()
    if not street:
        return None
    return "ul. " + street

def digits_only(s: object) -> str:
    if pd.isna(s):
        return ""
    return re.sub(r"\D", "", str(s))

def clean_price_numeric(price: object, year: object) -> Optional[int]:
    d = digits_only(price)
    if not d:
        return pd.NA
    if pd.notna(year):
        try:
            y = str(int(year))
            if len(y) == 2:
                y = "20"+y if int(y) < 50 else "19"+y
            if d.startswith(y):
                d = d[len(y):]
        except Exception:
            pass
    d = d.lstrip("0")
    if not d:
        return pd.NA
    try:
        return int(d)
    except Exception:
        try:
            return int(float(d))
        except Exception:
            return pd.NA

def mode1(s: pd.Series):
    s = s.dropna()
    if s.empty: return None
    m = s.mode()
    return m.iloc[0] if not m.empty else None

def unique1(s: pd.Series):
    s = s.dropna().unique()
    return s[0] if len(s) == 1 else None

# ------------------------ core filling ------------------------

def fill_from_internal(df: pd.DataFrame, W: str, P: str, G: str, M: str, D: str, U: str) -> int:
    map_m_to_w = df.dropna(subset=[M, W]).groupby(M)[W].apply(mode1).to_dict()
    map_p_to_w = df.dropna(subset=[P, W]).groupby(P)[W].apply(mode1).to_dict()
    map_g_to_w = df.dropna(subset=[G, W]).groupby(G)[W].apply(mode1).to_dict()

    map_m_to_p = df.dropna(subset=[M, P]).groupby(M)[P].apply(mode1).to_dict()
    map_m_to_g = df.dropna(subset=[M, G]).groupby(M)[G].apply(mode1).to_dict()
    map_mg_to_p = df.dropna(subset=[M, G, P]).groupby([M, G])[P].apply(mode1).to_dict()
    map_mp_to_g = df.dropna(subset=[M, P, G]).groupby([M, P])[G].apply(mode1).to_dict()

    map_md_to_u = df.dropna(subset=[M, D, U]).groupby([M, D])[U].apply(mode1).to_dict()
    map_m_to_d  = df.dropna(subset=[M, D]).groupby(M)[D].apply(mode1).to_dict()

    changes = 0
    for i, row in df.iterrows():
        m,p,g,w,dz,u = row[M], row[P], row[G], row[W], row[D], row[U]
        if pd.isna(w):
            cand = None
            if pd.notna(m) and m in map_m_to_w: cand = map_m_to_w[m]
            elif pd.notna(p) and p in map_p_to_w: cand = map_p_to_w[p]
            elif pd.notna(g) and g in map_g_to_w: cand = map_g_to_w[g]
            if cand is not None:
                df.at[i, W] = cand; changes += 1
        if pd.isna(p):
            cand = None
            if pd.notna(m) and m in map_m_to_p: cand = map_m_to_p[m]
            if cand is None and pd.notna(m) and pd.notna(g) and (m,g) in map_mg_to_p: cand = map_mg_to_p[(m,g)]
            if cand is not None:
                df.at[i, P] = cand; changes += 1
        if pd.isna(g):
            cand = None
            if pd.notna(m) and m in map_m_to_g: cand = map_m_to_g[m]
            if cand is None and pd.notna(m) and pd.notna(p) and (m,p) in map_mp_to_g: cand = map_mp_to_g[(m,p)]
            if cand is not None:
                df.at[i, G] = cand; changes += 1
        if pd.isna(dz) and pd.notna(m) and m in map_m_to_d:
            df.at[i, D] = map_m_to_d[m]; changes += 1
        if pd.isna(u) and pd.notna(m) and pd.notna(dz) and (m,dz) in map_md_to_u:
            df.at[i, U] = map_md_to_u[(m,dz)]; changes += 1
    return changes

def build_teryt_maps(teryt: pd.DataFrame, TW: Optional[str], TP: Optional[str], TG: Optional[str],
                     TM: Optional[str], TD: Optional[str], TU: Optional[str]) -> Tuple[dict,dict,dict,dict]:
    t = teryt.copy()
    t["_n_TM"] = t[TM].apply(norm_text) if TM else None
    t["_n_TD"] = t[TD].apply(norm_text) if TD else None
    t["_n_TU"] = t[TU].apply(norm_text) if TU else None
    t["_n_TW"] = t[TW].apply(norm_text) if TW else None
    t["_n_TP"] = t[TP].apply(norm_text) if TP else None
    t["_n_TG"] = t[TG].apply(norm_text) if TG else None

    def agg_map(keys: List[str]):
        grp = t.groupby(keys, dropna=False).agg(
            TW_val=("_n_TW", mode1),
            TP_val=("_n_TP", mode1),
            TG_val=("_n_TG", mode1),
            TU_one=("_n_TU", unique1),
        ).reset_index()
        return {tuple(row[k] for k in keys): (row.TW_val, row.TP_val, row.TG_val, row.TU_one)
                for _, row in grp.iterrows()}

    m_map  = agg_map(["_n_TM"])
    md_map = agg_map(["_n_TM","_n_TD"])
    mu_map = agg_map(["_n_TM","_n_TU"])
    mdu_map= agg_map(["_n_TM","_n_TD","_n_TU"])
    return m_map, md_map, mu_map, mdu_map

def fill_from_teryt(df: pd.DataFrame, teryt: pd.DataFrame,
                    W: str, P: str, G: str, M: str, D: str, U: str,
                    TW: Optional[str], TP: Optional[str], TG: Optional[str],
                    TM: Optional[str], TD: Optional[str], TU: Optional[str]) -> int:
    df["_n_m"] = df[M].apply(norm_text)
    needed = set(x for x in df["_n_m"].dropna().unique())
    t_small = teryt[teryt[TM].apply(norm_text).isin(needed)].copy() if TM else teryt.copy()

    m_map, md_map, mu_map, mdu_map = build_teryt_maps(t_small, TW, TP, TG, TM, TD, TU)

    df["_n_d"] = df[D].apply(norm_text)
    df["_n_u"] = df[U].apply(norm_text)

    def pick_from_maps(m, d, u):
        order = [((m,d,u), mdu_map), ((m,u), mu_map), ((m,d), md_map), ((m,), m_map)]
        for k, mp in order:
            if None in k:
                continue
            if k in mp:
                return mp[k]
        return (None,None,None,None)

    changes = 0
    for i, row in df.iterrows():
        m = row["_n_m"]; d = row["_n_d"]; u = row["_n_u"]
        Wv, Pv, Gv, Uv = pick_from_maps(m,d,u)
        updated = False
        if pd.isna(row[W]) and Wv is not None:
            df.at[i, W] = Wv; updated = True
        if pd.isna(row[P]) and Pv is not None:
            df.at[i, P] = Pv; updated = True
        if pd.isna(row[G]) and Gv is not None:
            df.at[i, G] = Gv; updated = True
        if pd.isna(row[U]) and Uv is not None:
            street = Uv
            df.at[i, U] = f"ul. {street}" if street and not str(street).startswith("ul") else street
            updated = True
        if updated: changes += 1

    df.drop(columns=[c for c in ["_n_m","_n_d","_n_u"] if c in df.columns], inplace=True)
    return changes

# ------------------------ processing + in-place write ------------------------

def find_source_sheet(xlsx: Path) -> str:
    xl = pd.ExcelFile(xlsx, engine="openpyxl")
    for nm in xl.sheet_names:
        tmp = xl.parse(nm, nrows=1)
        if "Nr KW" in tmp.columns:
            return nm
    return xl.sheet_names[0]

def write_df_inplace_no_new_sheets(xlsx: Path, sheet_name: str, df: pd.DataFrame) -> None:
    """Nadpisuje TYLKO istniejący arkusz, bez tworzenia nowych."""
    wb = load_workbook(xlsx)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb[wb.sheetnames[0]]

    # wyczyść arkusz
    ws.delete_rows(1, ws.max_row)

    # --- KLUCZ 1: usuń nieskończoności ---
    df_to_write = df.replace([np.inf, -np.inf], np.nan).copy()

    # --- KLUCZ 2: zbij dtype'y do object (żeby móc wstawić None zamiast <NA>) ---
    df_to_write = df_to_write.astype(object)

    # --- KLUCZ 3: zamień każde NA/NaN na None (openpyxl akceptuje None, ale nie <NA>) ---
    na_mask = pd.isna(df_to_write)
    df_to_write[na_mask] = None

    # nagłówki + dane
    for row in dataframe_to_rows(df_to_write, index=False, header=True):
        ws.append(row)

    wb.save(xlsx)

def process_inplace(input_path: Path, teryt_path: Optional[Path]) -> None:
    if not input_path.exists():
        raise FileNotFoundError(f"Nie znaleziono pliku wejściowego: {input_path}")

    sheet = find_source_sheet(input_path)
    df = pd.read_excel(input_path, sheet_name=sheet, engine="openpyxl")

    # wymagane kolumny + mapowanie
    cmap = ensure_columns(df, [
        "cena","rok_budowy","metry","cena_za_metr",
        "wojewodztwo","powiat","gmina","miejscowosc","dzielnica","ulica","link"
    ])
    CENA, ROK, METRY, CENAM2 = cmap["cena"], cmap["rok_budowy"], cmap["metry"], cmap["cena_za_metr"]
    W, P, G, M, D, U, LNK = (cmap["wojewodztwo"], cmap["powiat"], cmap["gmina"],
                              cmap["miejscowosc"], cmap["dzielnica"], cmap["ulica"], cmap["link"])

    # Tekstowe → object
    for c in [W, P, G, M, D, U, LNK]:
        if c in df.columns:
            df[c] = df[c].astype("object")

    # Liczbowe → numeric
    for c in [CENA, METRY, CENAM2, ROK]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    # 1) czyszczenie ceny → liczba
    df[CENA] = [clean_price_numeric(val, yr) for val, yr in zip(df[CENA], df[ROK])]
    try:
        df[CENA] = df[CENA].astype("Int64")
    except Exception:
        df[CENA] = pd.to_numeric(df[CENA], errors="coerce")

    # 2) cena za m2 (float) — metry mogą być 0/NaN → zabezpiecz inf
    cena_num  = pd.to_numeric(df[CENA], errors="coerce")
    metry_num = pd.to_numeric(df[METRY], errors="coerce")
    cena_za_m = cena_num / metry_num
    cena_za_m = cena_za_m.replace([np.inf, -np.inf], np.nan)
    df[CENAM2] = cena_za_m.round(2)

    # 3) ulica z linku (dla braków)
    miss_u = df[U].isna()
    df.loc[miss_u, U] = df.loc[miss_u, LNK].apply(extract_street_from_link)

    # 4) fill z własnych danych
    for _ in range(3):
        ch = fill_from_internal(df, W, P, G, M, D, U)
        if ch == 0:
            break

    # 5) fill z TERYT (opcjonalnie)
    if teryt_path and teryt_path.exists():
        teryt = pd.read_excel(teryt_path, engine="openpyxl")
        tmap = norm_colnames(teryt.columns)
        aliases = {
            "wojewodztwo": ["wojewodztwo","woj","nazwa_woj","nazwa województwa","województwo"],
            "powiat": ["powiat","nazwa_pow","nazwa powiatu"],
            "gmina": ["gmina","nazwa_gmi","nazwa gminy"],
            "miejscowosc": ["miejscowosc","miejscowość","nazwa miejscowosci","nazwa_miejscowosci","nazwa miejscowości"],
            "dzielnica": ["dzielnica","jednostka pomocnicza","jop","czesc miejscowosci","część miejscowości"],
            "ulica": ["ulica","nazwa ulicy","naz_ul","nazwa_ulicy"],
        }
        def pick(keys):
            for k in keys:
                if k in tmap: return tmap[k]
            return None
        TW = pick(aliases["wojewodztwo"]); TP = pick(aliases["powiat"]); TG = pick(aliases["gmina"])
        TM = pick(aliases["miejscowosc"]); TD = pick(aliases["dzielnica"]); TU = pick(aliases["ulica"])

        teryt["_n_TM_tmp"] = teryt[TM].apply(norm_text) if TM else None
        needed = set(x for x in df[M].apply(norm_text).dropna().unique())
        t_small = teryt[teryt["_n_TM_tmp"].isin(needed)].drop(columns=["_n_TM_tmp"])
        fill_from_teryt(df, t_small, W, P, G, M, D, U, TW, TP, TG, TM, TD, TU)

    # upewnij typy liczbowe po uzupełnieniach
    for c in [CENA, METRY, CENAM2, ROK]:
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")
    try:
        df[CENA] = df[CENA].astype("Int64")
    except Exception:
        pass

    # zapis IN-PLACE (bez tworzenia nowych arkuszy)
    write_df_inplace_no_new_sheets(input_path, sheet, df)

# ----------------------------- CLI -----------------------------

def main():
    ap = argparse.ArgumentParser(
        description="Czyści i uzupełnia dane (cena, adresy, TERYT) i nadpisuje ten sam arkusz — bez tworzenia nowych. Kolumny liczbowe zapisuje jako liczby."
    )
    ap.add_argument("--input", "-i", help="Wejściowy Excel. Domyślnie: ~/Desktop/baza danych/Baza danych.xlsx")
    ap.add_argument("--teryt", "-t", help="TERYT.xlsx (opcjonalnie). Domyślnie: ~/Desktop/baza danych/TERYT.xlsx")
    args = ap.parse_args()

    in_path = resolve_default(args.input, "Baza danych.xlsx")
    t_path  = resolve_default(args.teryt, "TERYT.xlsx")

    process_inplace(in_path, t_path if t_path.exists() else None)
    print(f"Nadpisano arkusz w pliku: {in_path}")

if __name__ == "__main__":
    main()
