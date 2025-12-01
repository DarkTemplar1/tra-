# -*- coding: utf-8 -*-
import sys
import csv
import argparse
from pathlib import Path
from typing import List
import tkinter as tk
from tkinter import filedialog, messagebox
from openpyxl import Workbook, load_workbook

# --- ustawienia arkusza raportu ---
RAPORT_SHEET = "raport"
RAPORT_ODF = "raport_odfiltrowane"

# kolumny „opisowe” – te które już widzisz na screenie
REQ_COLS: List[str] = [
    "Nr KW","Typ Księgi","Stan Księgi","Województwo","Powiat","Gmina",
    "Miejscowość","Dzielnica","Położenie","Nr działek po średniku","Obręb po średniku",
    "Ulica","Sposób korzystania","Obszar","Ulica(dla budynku)",
    "przeznaczenie (dla budynku)","Ulica(dla lokalu)","Nr budynku( dla lokalu)",
    "Przeznaczenie (dla lokalu)","Cały adres (dla lokalu)","Czy udziały?"
]

# NOWE kolumny do wyliczeń
VALUE_COLS: List[str] = [
    "Średnia cena za m2 ( z bazy)",
    "Średnia skorygowana cena za m2",
    "Statystyczna wartość nieruchomości",
]

# nagłówek CSV dla województw (tu bez zmian)
WYNIKI_HEADER: List[str] = [
    "cena","cena_za_metr","metry","liczba_pokoi","pietro","rynek","rok_budowy","material",
    "wojewodztwo","powiat","gmina","miejscowosc","dzielnica","ulica","link",
]

SUPPORTED = {".xlsx", ".xlsm"}

VOIVODESHIPS_LABEL_SLUG: list[tuple[str, str]] = [
    ("Dolnośląskie", "dolnoslaskie"),
    ("Kujawsko-Pomorskie", "kujawsko-pomorskie"),
    ("Lubelskie", "lubelskie"),
    ("Lubuskie", "lubuskie"),
    ("Łódzkie", "lodzkie"),
    ("Małopolskie", "malopolskie"),
    ("Mazowieckie", "mazowieckie"),
    ("Opolskie", "opolskie"),
    ("Podkarpackie", "podkarpackie"),
    ("Podlaskie", "podlaskie"),
    ("Pomorskie", "pomorskie"),
    ("Śląskie", "slaskie"),
    ("Świętokrzyskie", "swietokrzyskie"),
    ("Warmińsko-Mazurskie", "warminsko-mazurskie"),
    ("Wielkopolskie", "wielkopolskie"),
    ("Zachodniopomorskie", "zachodniopomorskie"),
]

# --------------------- Desktop/Pulpit ---------------------
def _detect_desktop() -> Path:
    home = Path.home()
    for name in ("Desktop", "Pulpit"):
        p = home / name
        if p.exists():
            return p
    return home

def ensure_base_dirs(base_override: Path | None = None) -> Path:
    """
    Zwraca katalog bazowy 'baza danych' i upewnia się, że istnieją:
      • <base>/linki
      • <base>/województwa
      • <base>/timing.csv (z nagłówkiem)
    """
    if base_override:
        base = Path(base_override)
    else:
        base = _detect_desktop() / "baza danych"

    (base / "linki").mkdir(parents=True, exist_ok=True)
    (base / "województwa").mkdir(parents=True, exist_ok=True)

    timing = base / "timing.csv"
    if not timing.exists():
        with timing.open("w", encoding="utf-8-sig", newline="") as f:
            w = csv.writer(f)
            w.writerow(["Województwo", "Stan pobierania"])
    return base
# ---------------------------------------------------------

# --------------------- CSV helpery -----------------------
def _ensure_csv(path: Path, header: List[str]) -> bool:
    if path.exists():
        return False
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.writer(f)
        if header:
            w.writerow(header)
    return True

def create_voivodeship_csvs(base: Path) -> dict:
    created = {"linki": 0, "województwa": 0}
    linki_dir = base / "linki"
    woj_dir = base / "województwa"
    for label, _slug in VOIVODESHIPS_LABEL_SLUG:
        if _ensure_csv(linki_dir / f"{label}.csv", ["link"]):
            created["linki"] += 1
        if _ensure_csv(woj_dir / f"{label}.csv", WYNIKI_HEADER):
            created["województwa"] += 1
    return created
# ---------------------------------------------------------

# ----------------- dopisywanie kolumn w raporcie -----------------
def ensure_report_columns(xlsx: Path) -> None:
    """
    Dopisuje do pliku Excela (raport) brakujące kolumny z REQ_COLS i VALUE_COLS.
    Jeśli arkusz 'raport' nie istnieje, używa pierwszego arkusza.
    """
    xlsx = xlsx.expanduser()
    if not xlsx.exists():
        raise FileNotFoundError(f"Nie znaleziono pliku: {xlsx}")
    if xlsx.suffix.lower() not in SUPPORTED:
        raise ValueError(f"Obsługiwane tylko pliki Excel: {SUPPORTED} (podano: {xlsx.suffix})")

    wb = load_workbook(xlsx)
    if RAPORT_SHEET in wb.sheetnames:
        ws = wb[RAPORT_SHEET]
    else:
        ws = wb[wb.sheetnames[0]]

    # aktualne nagłówki (pierwszy wiersz)
    headers: list[str] = []
    for cell in ws[1]:
        val = cell.value
        headers.append("" if val is None else str(val))

    # lista kolumn do upewnienia się
    target_cols = REQ_COLS + VALUE_COLS

    for name in target_cols:
        if name not in headers:
            col_idx = len(headers) + 1  # kolejna wolna kolumna
            ws.cell(row=1, column=col_idx).value = name
            headers.append(name)

    wb.save(xlsx)

# ---------------------------------------------------------

def _gui_pick_and_add_columns():
    """Proste okienko: wybierz plik Excela i dopisz kolumny."""
    root = tk.Tk()
    root.withdraw()
    path = filedialog.askopenfilename(
        title="Wybierz plik raportu (Excel)",
        filetypes=[("Excel", "*.xlsx;*.xlsm"), ("Wszystkie pliki", "*.*")]
    )
    if not path:
        return
    try:
        ensure_report_columns(Path(path))
        messagebox.showinfo("Kolumny", f"Dopisano brakujące kolumny w pliku:\n{path}")
    except Exception as e:
        messagebox.showerror("Błąd", f"Nie udało się dopisać kolumn:\n{e}")

# ----------------------------- CLI -----------------------------
def main():
    parser = argparse.ArgumentParser(
        description="PriceBot – tworzenie struktury 'baza danych' lub dopisywanie kolumn w raporcie."
    )
    parser.add_argument("--base-dir", help="Gdzie utworzyć 'baza danych' (domyślnie: Desktop/Pulpit).")
    parser.add_argument("--in", dest="inp", type=Path,
                        help="Plik raportu (Excel), do którego dopisać brakujące kolumny.")
    args = parser.parse_args()

    # 1) jeśli podano --in → tryb dopisywania kolumn
    if args.inp:
        try:
            ensure_report_columns(args.inp)
            print(f"[kolumny] Dodano brakujące kolumny w pliku: {args.inp}")
        except Exception as e:
            print(f"[ERR] Nie udało się dopisać kolumn: {e}", file=sys.stderr)
            sys.exit(1)
        return

    # 2) w przeciwnym razie – stare zachowanie: struktura 'baza danych'
    base_override = Path(args.base_dir) if args.base_dir else None
    base = ensure_base_dirs(base_override)
    created = create_voivodeship_csvs(base)

    print(f"[kolumny] Baza: {base}")
    print(f"[kolumny] Utworzone: linki={created['linki']}, województwa={created['województwa']}")

if __name__ == "__main__":
    # jeśli uruchomisz podwójnym kliknięciem bez argumentów,
    # otworzy się okienko wyboru pliku i dopisania kolumn
    if len(sys.argv) == 1:
        _gui_pick_and_add_columns()
    else:
        main()
